import sys
import time
import openpyxl
from backward import BackwardChaining
from forward import ForwardChaining
from Class import *
from Class import ConvertData

# 1. thu thập thông tin người dùng
def thong_tin_user():
    print("*Bot: Xin chào, tôi là chatbot chuẩn đoán các bệnh về dinh dưỡng ở người cao tuổi!")
    time.sleep(1)
    print("*Bot: Để nhận lời khuyên và chuẩn đoán chi tiết, hãy để lại email, tên và số điện thoại của bạn")
    time.sleep(1)
    print("*Bot: Hãy nhập tên của bạn")
    i = input("*User: ")
    user.name = validate.validate_name(i)
    time.sleep(1)
    print("*Bot: Hãy nhập email của bạn")
    i = input("*User: ")
    user.email = validate.validate_email(i)
    time.sleep(1)
    print("*Bot: Hãy nhập số điện thoại của bạn")
    i = input("*User: ")
    user.phoneNumber = validate.validate_phonenumber(i)
    time.sleep(1)
    print(f'*Bot: Cảm ơn bạn đã nhập thông tin. Thông tin của bạn là: {user.name}, {user.email}, {user.phoneNumber}')
    time.sleep(1)
    return user

# 2. thu thập một số triệu chứng phổ biến của bệnh dinh dưỡng mà người dùng mắc phải
def trieuchung_phobien(trieuchung_cua_user, user):
    All_tc_list = [db.trieuchung[4], db.trieuchung[17],
                 db.trieuchung[24], db.trieuchung[46]]
    print(f'*Bot: Bây giờ, tôi muốn hỏi {user.name} về một số triệu chứng phổ biến thường gặp trong bệnh về dinh dưỡng ở người cao tuổi.')
    New_All_tc_list = []
    for i in All_tc_list:
        New_All_tc_list.append(i["idTrieuChung"])
    check = [1,2,3,4]
    while (1):
        if (len(trieuchung_cua_user) == len(All_tc_list)):
            break
        if (len(trieuchung_cua_user) == 0):
            time.sleep(1)
            print(f'*Bot: {user.name} hãy nhập các triệu chứng mình gặp phải bằng cách nhập số thứ tự của triệu chứng nhé.')
            
        else:
            time.sleep(1)
            print(f'*Bot: {user.name} có còn gặp triệu trứng nào nữa ở dưới nữa không?')

        
        for i in range(len(All_tc_list)):
            if check[i]!=0:
                print(f'      {check[i]}.{All_tc_list[i]["noiDung"]}')
            
        print("      0.Tôi không có triệu chứng nào ở trên.")
        i = input("*User: ")
        answer = validate.validate_input_number_form(i)
        
        if (answer == '0'):
            break
        elif (int(answer) < 0 or int(answer) > 4):
            print('*Bot: Bạn vui lòng nhập 1 số từ 0 tới 4')
            continue
        else:
            check[int(answer)-1] =0
            trieuchung_cua_user.append(All_tc_list[int(answer)-1])
    time.sleep(1)
    print( f'*Bot: Danh sách các triệu chứng {user.name} đang gặp phải:',end=' ')
    print([i['noiDung'] for i in trieuchung_cua_user],end=', ') 
    print()   
    return trieuchung_cua_user

# 3. thu thập các triệu chứng chi tiết của mệt mỏi mà người dùng mắc phải
def trieuchung_ct_metmoi(trieuchung_cua_user, user):
    trieuchung_metmoi = [
        db.trieuchung[5],
        db.trieuchung[9],
        db.trieuchung[15],
        db.trieuchung[27],
        db.trieuchung[32],
        db.trieuchung[33]
    ]
    
    check = {'idTrieuChung': 'S05', 'noiDung': 'Mệt mỏi'}
    ok=0
    check1 = [1,2,3,4,5,6]
    if (check in trieuchung_cua_user):
        time.sleep(1)
        print( f'*Bot: Tiếp theo tôi muốn biết chi tiết hơn về tình trạng mệt mỏi của {user.name}.')
        while (1):
            if ok==0:
                time.sleep(1)
                print(f'*Bot: {user.name} hãy nhập các triệu chứng mình gặp phải bằng cách nhập số thứ tự của triệu chứng nhé.')
            else:
                time.sleep(1)
                print(f'*Bot: {user.name} có còn gặp triệu trứng nào nữa ở dưới nữa không?')
            
            for i in range(len(trieuchung_metmoi)):
                if check1[i]!=0:
                    print(f'      {check1[i]}.{trieuchung_metmoi[i]["noiDung"]}')
            
            print('      0.Bỏ qua')
            i = input('*User: ')
            answer = validate.validate_input_number_form(i)

            if (int(answer) < 0 or int(answer) > 6):
                print('*Bot: Bạn vui lòng nhập 1 số từ 0 tới 6')
                continue
            elif (answer == '0'):
                break
            else:
                trieuchung_cua_user.append(trieuchung_metmoi[int(answer)-1])
                check1[int(answer)-1] =0
                ok=1
        time.sleep(1)
        print( f'*Bot: Danh sách các triệu chứng {user.name} đang gặp phải:',end=' ')
        print([i['noiDung'] for i in trieuchung_cua_user],end=', ') 
        print()   
    
    return trieuchung_cua_user

# 4. Cthu thập các triệu chứng chi tiết của suy giảm thị lực mà người dùng mắc phải
def trieuchung_ct_thiluc(trieuchung_cua_user, user):
    trieuchung_thiluc = [
        db.trieuchung[44],
        db.trieuchung[45]
    ]
    
    check = {'idTrieuChung': 'S25', 'noiDung': 'suy giảm thị lực'}
    ok=0
    check1 = [1,2]
    if (check in trieuchung_cua_user):
        time.sleep(1)
        print( f'*Bot: Tiếp theo tôi muốn biết chi tiết hơn về tình trạng suy giảm thị lực của {user.name}.')
        while (1):
            if ok==0:
                time.sleep(1)
                print(f'*Bot: {user.name} hãy nhập các triệu chứng mình gặp phải bằng cách nhập số thứ tự của triệu chứng nhé.')
            else:
                time.sleep(1)
                print(f'*Bot: {user.name} có còn gặp triệu trứng nào nữa ở dưới nữa không?')
            
            for i in range(len(trieuchung_thiluc)):
                if check1[i]!=0:
                    print(f'      {check1[i]}.{trieuchung_thiluc[i]["noiDung"]}')
            
            print('      0.Bỏ qua')
            i = input('*User: ')
            answer = validate.validate_input_number_form(i)

            if (int(answer) < 0 or int(answer) > 2):
                print('*Bot: Bạn vui lòng nhập 1 số từ 0 tới 2')
                continue
            elif (answer == '0'):
                break
            else:
                trieuchung_cua_user.append(trieuchung_thiluc[int(answer)-1])
                check1[int(answer)-1] =0
                ok=1
        time.sleep(1)
        print( f'*Bot: Danh sách các triệu chứng {user.name} đang gặp phải:',end=' ')
        print([i['noiDung'] for i in trieuchung_cua_user],end=', ') 
        print()   
    
    return trieuchung_cua_user


# 5. thu thập các triệu chứng chi tiết của đau mỏi xương khớp mà người dùng mắc phải
def trieuchung_ct_xuongkhop(trieuchung_cua_user, user):
    trieuchung_xuongkhop = [
        db.trieuchung[18],
        db.trieuchung[19],
        db.trieuchung[20],
        db.trieuchung[21],
        db.trieuchung[37],
        db.trieuchung[38]
    ]
    
    check = {'idTrieuChung': 'S18', 'noiDung': 'đau mỏi xương khớp'}
    ok=0
    check1 = [1,2,3,4,5,6]
    if (check in trieuchung_cua_user):
        time.sleep(1)
        print( f'*Bot: Tiếp theo tôi muốn biết chi tiết hơn về tình trạng đau mỏi xương khớp của {user.name}.')
        while (1):
            if ok==0:
                time.sleep(1)
                print(f'*Bot: {user.name} hãy nhập các triệu chứng mình gặp phải bằng cách nhập số thứ tự của triệu chứng nhé.')
            else:
                time.sleep(1)
                print(f'*Bot: {user.name} có còn gặp triệu trứng nào nữa ở dưới nữa không?')
            
            for i in range(len(trieuchung_xuongkhop)):
                if check1[i]!=0:
                    print(f'      {check1[i]}.{trieuchung_xuongkhop[i]["noiDung"]}')
            
            print('      0.Bỏ qua')
            i = input('*User: ')
            answer = validate.validate_input_number_form(i)

            if (int(answer) < 0 or int(answer) > 6):
                print('*Bot: Bạn vui lòng nhập 1 số từ 0 tới 6')
                continue
            elif (answer == '0'):
                break
            else:
                trieuchung_cua_user.append(trieuchung_xuongkhop[int(answer)-1])
                check1[int(answer)-1] =0
                ok=1
        time.sleep(1)
        print( f'*Bot: Danh sách các triệu chứng {user.name} đang gặp phải:',end=' ')
        print([i['noiDung'] for i in trieuchung_cua_user],end=', ') 
        print()   
    
    return trieuchung_cua_user

# 6. thu thập các triệu chứng chi tiết của cân nặng thay đổi thất thường mà người dùng mắc phải
def trieuchung_ct_cannang(trieuchung_cua_user, user):
    trieuchung_cannang = [
        db.trieuchung[13],
        db.trieuchung[28]
    ]
    
    check = {'idTrieuChung': 'S47', 'noiDung': 'cân nặng thay đổi thất thường'}
    ok=0
    check1 = [1,2]
    if (check in trieuchung_cua_user):
        time.sleep(1)
        print( f'*Bot: Tiếp theo tôi muốn biết chi tiết hơn về tình trạng cân nặng thất thường của {user.name}.')
        while (1):
            if ok==0:
                time.sleep(1)
                print(f'*Bot: {user.name} hãy nhập các triệu chứng mình gặp phải bằng cách nhập số thứ tự của triệu chứng nhé.')
            else:
                time.sleep(1)
                print(f'*Bot: {user.name} có còn gặp triệu trứng nào nữa ở dưới nữa không?')
            
            for i in range(len(trieuchung_cannang)):
                if check1[i]!=0:
                    print(f'      {check1[i]}.{trieuchung_cannang[i]["noiDung"]}')
            
            print('      0.Bỏ qua')
            i = input('*User: ')
            answer = validate.validate_input_number_form(i)

            if (int(answer) < 0 or int(answer) > 2):
                print('*Bot: Bạn vui lòng nhập 1 số từ 0 tới 2')
                continue
            elif (answer == '0'):
                break
            else:
                trieuchung_cua_user.append(trieuchung_cannang[int(answer)-1])
                check1[int(answer)-1] =0
                ok=1
        time.sleep(1)
        print( f'*Bot: Danh sách các triệu chứng {user.name} đang gặp phải:',end=' ')
        print([i['noiDung'] for i in trieuchung_cua_user],end=', ') 
        print()   
    
    return trieuchung_cua_user

# 7. sử dụng suy diễn tiến để dự đoán các bệnh người dùng có thể gặp phải
def forward_chaining(rule, fact, goal, file_name,user):
    fc = ForwardChaining(rule, fact, None, file_name)

    list_predicted_disease = [i for i in fc.facts if i[0] == "D"]
    time.sleep(1)
    print(f'*Bot: Dựa trên những triệu chứng trên, tôi dự đoán {user.name} có thể bị bệnh :', end=" ")
    for i in list_predicted_disease:
        temp = db.get_benh_by_id(i)
        print(temp['tenBenh'], end=', ')
    print()
    time.sleep(1)
    print(f'*Bot: Trên đây là chuẩn đoán sơ bộ của tôi. Tiếp theo, tôi sẽ hỏi {user.name} một số câu hỏi để đưa ra kết quả chính xác.')
    return list_predicted_disease


# 8. sử dụng suy diễn lùi để đưa ra suy đoán chính xác bệnh mà người dùng mắc phải
def backward_chaining(suy_dien_lui,trieuchung_cua_user_id,list_predicted_disease,file_name ):
    predictD=list_predicted_disease
    rule=suy_dien_lui
    all_rule=db.gettrieuchung()
    fact_real=trieuchung_cua_user_id
    benh=0
    for g in predictD:
        goal=g
        D=db.get_benh_by_id(goal) #Chứa thông tin của bệnh có id == goal
        time.sleep(1)
        print(f"*Bot: Theo như các triệu chứng ban đầu và có thể bạn mắc bệnh {D['tenBenh']}, sau đây chúng tôi muốn hỏi bạn một vài câu hỏi để tìm hiểu về bệnh bạn đang mắc phải")
        all_s_in_D=all_rule[goal]
        all_s_in_D=sorted(set(all_s_in_D)-set(fact_real))
        d=searchindexrule(rule,goal)
        
        b=BackwardChaining(rule,fact_real,goal,file_name) # kết luận trong trường hợp các luât trước đã suy ra đk luôn
        
        if b.result1==True:# đoạn đầu
            time.sleep(1)
            print("*Bot: Bạn mắc bệnh {}".format(D['tenBenh']))
            time.sleep(1)
            D['moTa']=D['moTa'].replace("/n","\n      ")
            print(f"      {D['moTa']}")
            time.sleep(1)
            print(f"*Bot: Nguyên nhân của bệnh là:")
            D['nguyenNhan']=D['nguyenNhan'].replace("/n","\n      ")
            print(f"      {D['nguyenNhan']}")
            time.sleep(1)
            print(f"*Bot: Sau đây là một số lời khuyên tôi dành cho bạn.")
            D['loiKhuyen']=D['loiKhuyen'].replace("/n","\n     ")
            print(f"     {D['loiKhuyen']}")
            print("*Bot: Cám ơn bạn đã sử dụng chatbot")
            return goal,fact_real
        
        while(len(all_s_in_D)>0):
            s=db.get_trieuchung_by_id(all_s_in_D[0])
            time.sleep(1)
            question=f"*Bot: Bạn có bị triệu chứng {s['noiDung']} không?"
            print(question)
            i = input('*User: ')
            answer = validate.validate_binary_answer(i)
            
            if answer== True :
                fact_real.append(all_s_in_D[0])
                b=BackwardChaining(rule,fact_real,goal,file_name)
                list_no_result,lsD=get_s_in_d(all_s_in_D[0],goal,rule,d,1)
                d=sorted(set(d)-set(lsD))
                all_s_in_D=sorted(set(list_no_result)-set(fact_real))
                if b.result1==True:
                    benh=1
                    break
            if answer==False :
                list_no_result,lsD=get_s_in_d(all_s_in_D[0],goal,rule,d,0) 
                d=sorted(set(d)-set(lsD))
                all_s_in_D=sorted(set(list_no_result)-set(fact_real))
            if len(d)==0: 
                time.sleep(1)
                print(f"*Bot: Có vẻ như bạn không mắc bệnh {D['tenBenh']}")
                break
        if benh==1:
            time.sleep(1)
            print("*Bot: Bạn mắc bệnh {}".format(D['tenBenh']))
            time.sleep(1)
            D['moTa']=D['moTa'].replace("/n","\n     ")
            print(f"      {D['moTa']}")
            time.sleep(1)
            print(f"*Bot: Nguyên nhân của bệnh là:")
            D['nguyenNhan']=D['nguyenNhan'].replace("/n","\n     ")
            print(f"      {D['nguyenNhan']}")
            time.sleep(1)
            print(f"*Bot: Sau đây là một số lời khuyên tôi dành cho bạn.")
            D['loiKhuyen']=D['loiKhuyen'].replace("/n","\n     ")
            print(f"     {D['loiKhuyen']}")
            print("*Bot: Cám ơn bạn đã sử dụng chatbot")
            
            return goal,fact_real
            break
    if benh==0:
        time.sleep(1)
        print(f"*Bot: Bạn không bị bệnh nào cả. Cám ơn bạn đã sử dụng ChatBot")
        return None, fact_real   

#9. lưu thông tin data người dùng vào tệp
def luu_thong_tin(benh_user, trieuchung_cua_user_id,user):
    tc = "'"
    for i in trieuchung_cua_user_id:
        tc+= '- '+ db.get_trieuchung_by_id(i)['noiDung']+'\n'
    
    bdd = db.get_benh_by_id(benh_user)['tenBenh']

    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb['Sheet1']
    l = len(list(sheet.values))
    sheet.cell(row=l+1, column=1,value=user.name)
    sheet.cell(row=l+1, column=2,value=user.phoneNumber)
    sheet.cell(row=l+1, column=3,value=user.email)
    sheet.cell(row=l+1, column=4,value=tc)
    sheet.cell(row=l+1, column=5,value=bdd)
    wb.save('data.xlsx')
# biến khởi tạo
user = User(None, None, None)
validate = Validate()
trieuchung_cua_user = []  # list các triệu chứng người dùng khi trả lời là yes

#kết nối database
db = ConvertData()
db.convertbenh()  # bang benh
db.converttrieuchung()  # bang trieu chung
db.getfc()
db.getbc()
suy_dien_lui = db.groupbc()
suy_dien_tien = db.groupfc()

user = thong_tin_user()
trieuchung_cua_user = []  # list các đối tượng triệu chứng

trieuchung_cua_user = trieuchung_phobien(trieuchung_cua_user, user)

trieuchung_cua_user = trieuchung_ct_metmoi(trieuchung_cua_user, user)

trieuchung_cua_user = trieuchung_ct_thiluc(trieuchung_cua_user, user)

trieuchung_cua_user = trieuchung_ct_xuongkhop(trieuchung_cua_user, user)

trieuchung_cua_user_id = [i['idTrieuChung'] for i in trieuchung_cua_user]
trieuchung_cua_user_id = list(set(trieuchung_cua_user_id))
trieuchung_cua_user_id.sort()

benh_dudoan_fc = forward_chaining(suy_dien_tien, trieuchung_cua_user_id, None, 'ex', user)

if len(benh_dudoan_fc)==0 :
    print("*Bot: Bạn không có dấu hiệu của bệnh nào cả.Cám ơn bạn đã sử dụng ChatBot")
    sys.exit()

benh_user,trieuchung_cua_user_id= backward_chaining(suy_dien_lui,trieuchung_cua_user_id,benh_dudoan_fc,"ex")
luu_thong_tin(benh_user, trieuchung_cua_user_id,user)
